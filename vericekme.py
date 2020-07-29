from __future__ import division
from openpyxl import Workbook,load_workbook
from bs4 import BeautifulSoup
import urllib
from urllib.request import urlopen

wb = load_workbook("sayisalUniVerileri.xlsx")
ws = wb.active
query="INSERT INTO netsayilari (bolumID, aytBiyoloji,aytKimya,aytMatematik,tytFen,tytSosyal,tytMatematik,tytTurkce) VALUES"
for satir in range(1,ws.max_row+1):
    bolumid = str(ws.cell(satir,1).value).replace('\t', '')
    url = "https://yokatlas.yok.gov.tr/content/lisans-dynamic/1210a.php?y="+bolumid
    sayfa = urllib.request.urlopen(url)
    soup = BeautifulSoup(sayfa, "html.parser")
    ana = soup.find('tbody')
    alt=ana.findAll('td',attrs={"class":"text-center"})
    durum=str(satir)+" - "+str(ws.max_row)
    print(durum)
    word =""
    counter=0
    for i in range(len(alt)):
        if(alt[i].text=="---"):
            continue
        if(i<=2):
            continue
        if(i%2==0):
            word+=alt[i].text.replace(',','.')+","
            counter+=1
            continue
    if(counter==8):
        query+="("
        query+='"'+bolumid+'"'+","
        query+=word+"),"
        print("eklendi")
    else:
        print("eklenmedi")

query = query.replace(',)',')')
query+=";"
query = query.replace(',;',';')
dosya = open('bolum-ve-ort-netler.sql', 'w')
dosya.write(query)
dosya.close()
